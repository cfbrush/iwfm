# test_xls.py
# Top-level tests for iwfm.xls package
# Copyright (C) 2026 University of California
# -----------------------------------------------------------------------------
# This information is free; you can redistribute it and/or modify it
# under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# This work is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# For a copy of the GNU General Public License, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301, USA.
# -----------------------------------------------------------------------------

import os
import pytest
from datetime import datetime


class TestPackagePublicAPI:
    """Public API exposed by iwfm.xls (backend-agnostic)."""

    def test_get_backend(self):
        from iwfm.xls import get_backend
        assert get_backend() in ('openpyxl', 'win32com')

    @pytest.mark.parametrize('name', [
        'create_workbook', 'open_workbook', 'save_workbook', 'close_workbook',
        'add_worksheet', 'get_worksheet', 'write_cells', 'write_budget_data',
    ])
    def test_new_api_exposed(self, name):
        """New (non-deprecated) API symbols are importable from iwfm.xls."""
        import iwfm.xls
        assert hasattr(iwfm.xls, name), f'iwfm.xls is missing {name}'
        assert callable(getattr(iwfm.xls, name))

    @pytest.mark.parametrize('name', [
        'excel_init', 'excel_new_workbook', 'excel_kill',
        'xl_open', 'xl_save', 'xl_quit', 'xl_write_2d',
        'bud2xl', 'write_budget_to_xl', 'write_2_excel',
    ])
    def test_legacy_api_still_importable(self, name):
        """Deprecated symbols remain importable for back-compat."""
        import iwfm.xls
        assert hasattr(iwfm.xls, name), f'iwfm.xls is missing legacy {name}'
        assert callable(getattr(iwfm.xls, name))


class TestDeprecationWarnings:
    """Each legacy function fires DeprecationWarning when called."""

    def test_bud2xl_deprecated(self, tmp_path):
        from iwfm.xls.bud2xl import bud2xl
        bogus_bud = tmp_path / "missing.bud"
        bogus_xls = tmp_path / "missing.xlsx"
        with pytest.warns(DeprecationWarning):
            try:
                bud2xl(str(bogus_bud), str(bogus_xls))
            except (FileNotFoundError, OSError, IndexError, SystemExit, ValueError):
                pass  # underlying op fails after warning fires; we only assert the warning

    def test_write_budget_to_xl_deprecated(self):
        from iwfm.xls.write_budget_to_xl import write_budget_to_xl
        from iwfm.xls import create_workbook
        wb = create_workbook()
        # minimal valid budget_data
        budget_data = [
            ['L1'], [['Date', 'V']],
            [__import__('pandas').DataFrame({'Date': [datetime(2020, 1, 1)], 'V': [1.0]})],
            [['T1', 'T2', 'T3']],
        ]
        with pytest.warns(DeprecationWarning):
            write_budget_to_xl(wb, budget_data)


class TestWrite2Excel:
    """write_2_excel produces an .xlsx file via openpyxl backend."""

    def test_write_2_excel_creates_file(self, tmp_path):
        from iwfm.xls import write_2_excel

        old_cwd = os.getcwd()
        os.chdir(tmp_path)
        try:
            data = [
                [[10.0, 20.0], [30.0, 40.0], [50.0, 60.0]],
                [[11.0, 21.0], [31.0, 41.0], [51.0, 61.0]],
            ]
            elements = 3
            time_steps = 2
            dates = [datetime(2020, 1, 1), datetime(2020, 2, 1)]
            file_base = "test_output"

            with pytest.warns(DeprecationWarning):
                write_2_excel(file_base, data, sheets=2, elements=elements,
                              time_steps=time_steps, dates=dates,
                              data_type='Test')

            output = tmp_path / "test_output.xlsx"
            assert output.exists()

            import openpyxl
            wb = openpyxl.load_workbook(output)
            try:
                # at least the two added sheets are present (default 'Sheet'
                # may also linger depending on backend cleanup)
                assert 'Test1' in wb.sheetnames
                assert 'Test2' in wb.sheetnames
                ws1 = wb['Test1']
                assert ws1['A1'].value == 'Test1'
                assert ws1['A2'].value == 'WYr'
            finally:
                wb.close()
        finally:
            os.chdir(old_cwd)


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
