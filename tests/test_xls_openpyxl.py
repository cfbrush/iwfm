# test_xls_openpyxl.py
# Unit tests for openpyxl backend
# Copyright (C) 2026 University of California
# -----------------------------------------------------------------------------
# This information is free; you can redistribute it and/or modify it
# under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# This work is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# For a copy of the GNU General Public License, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301, USA.
# -----------------------------------------------------------------------------

import pytest
from datetime import datetime

import pandas as pd


# ---------------------------------------------------------------------------
# Task 2: TestCreateWorkbook
# ---------------------------------------------------------------------------

class TestCreateWorkbook:
    """Tests for create_workbook function."""

    def test_create_workbook_returns_workbook(self):
        """create_workbook returns a workbook object."""
        from iwfm.xls._openpyxl_backend import create_workbook
        wb = create_workbook()
        assert wb is not None
        assert hasattr(wb, 'active')

    def test_create_workbook_stores_filename(self):
        """create_workbook stores filename for later save."""
        from iwfm.xls._openpyxl_backend import create_workbook
        wb = create_workbook(filename='test.xlsx')
        assert wb._iwfm_filename == 'test.xlsx'

    def test_create_workbook_without_filename(self):
        """create_workbook without filename stores None."""
        from iwfm.xls._openpyxl_backend import create_workbook
        wb = create_workbook()
        assert wb._iwfm_filename is None


# ---------------------------------------------------------------------------
# Task 3: TestOpenWorkbook (+ save / close)
# ---------------------------------------------------------------------------

class TestOpenWorkbook:
    """Tests for open_workbook / save_workbook / close_workbook."""

    def test_open_workbook_returns_workbook(self, tmp_path):
        """open_workbook returns a workbook object."""
        from iwfm.xls._openpyxl_backend import (
            create_workbook, save_workbook, open_workbook,
        )
        filepath = tmp_path / "test.xlsx"
        wb = create_workbook(str(filepath))
        save_workbook(wb)

        wb2 = open_workbook(str(filepath))
        assert wb2 is not None
        assert hasattr(wb2, 'active')

    def test_open_workbook_stores_filename(self, tmp_path):
        """open_workbook stores filename on workbook."""
        from iwfm.xls._openpyxl_backend import (
            create_workbook, save_workbook, open_workbook,
        )
        filepath = tmp_path / "test.xlsx"
        wb = create_workbook(str(filepath))
        save_workbook(wb)

        wb2 = open_workbook(str(filepath))
        assert wb2._iwfm_filename == str(filepath)

    def test_open_workbook_file_not_found(self):
        """open_workbook raises FileNotFoundError for missing file."""
        from iwfm.xls._openpyxl_backend import open_workbook
        with pytest.raises(FileNotFoundError):
            open_workbook('nonexistent.xlsx')

    def test_save_workbook_no_filename_raises(self):
        """save_workbook without filename + no associated filename → ValueError."""
        from iwfm.xls._openpyxl_backend import create_workbook, save_workbook
        wb = create_workbook()  # no filename
        with pytest.raises(ValueError):
            save_workbook(wb)

    def test_close_workbook_runs(self, tmp_path):
        """close_workbook completes without raising."""
        from iwfm.xls._openpyxl_backend import (
            create_workbook, save_workbook, close_workbook,
        )
        filepath = tmp_path / "test.xlsx"
        wb = create_workbook(str(filepath))
        save_workbook(wb)
        close_workbook(wb)  # no exception expected


# ---------------------------------------------------------------------------
# Task 4: TestWorksheetOperations
# ---------------------------------------------------------------------------

class TestWorksheetOperations:
    """Tests for add_worksheet / get_worksheet / write_cells."""

    def test_add_worksheet_creates_sheet(self):
        from iwfm.xls._openpyxl_backend import create_workbook, add_worksheet
        wb = create_workbook()
        ws = add_worksheet(wb, name='TestSheet')
        assert ws is not None
        assert ws.title == 'TestSheet'
        assert 'TestSheet' in wb.sheetnames

    def test_add_worksheet_without_name(self):
        from iwfm.xls._openpyxl_backend import create_workbook, add_worksheet
        wb = create_workbook()
        ws = add_worksheet(wb)
        assert ws is not None
        assert ws.title is not None

    def test_get_worksheet_by_name(self):
        from iwfm.xls._openpyxl_backend import (
            create_workbook, add_worksheet, get_worksheet,
        )
        wb = create_workbook()
        add_worksheet(wb, name='MySheet')
        ws = get_worksheet(wb, 'MySheet')
        assert ws.title == 'MySheet'

    def test_get_worksheet_by_index(self):
        from iwfm.xls._openpyxl_backend import create_workbook, get_worksheet
        wb = create_workbook()
        ws = get_worksheet(wb, 0)
        assert ws is not None

    def test_write_cells_writes_data(self):
        from iwfm.xls._openpyxl_backend import (
            create_workbook, get_worksheet, write_cells,
        )
        wb = create_workbook()
        ws = get_worksheet(wb, 0)
        data = [[1, 2, 3], [4, 5, 6]]
        write_cells(ws, data, start_row=1, start_col=1)
        assert ws.cell(row=1, column=1).value == 1
        assert ws.cell(row=1, column=3).value == 3
        assert ws.cell(row=2, column=1).value == 4
        assert ws.cell(row=2, column=3).value == 6

    def test_write_cells_with_offset(self):
        from iwfm.xls._openpyxl_backend import (
            create_workbook, get_worksheet, write_cells,
        )
        wb = create_workbook()
        ws = get_worksheet(wb, 0)
        write_cells(ws, [[10, 20]], start_row=5, start_col=3)
        assert ws.cell(row=5, column=3).value == 10
        assert ws.cell(row=5, column=4).value == 20

    def test_write_cells_empty_data(self):
        """Empty data list returns without raising."""
        from iwfm.xls._openpyxl_backend import (
            create_workbook, get_worksheet, write_cells,
        )
        wb = create_workbook()
        ws = get_worksheet(wb, 0)
        write_cells(ws, [])  # no-op


# ---------------------------------------------------------------------------
# Task 5: TestWriteBudgetData
# ---------------------------------------------------------------------------

class TestWriteBudgetData:
    """Tests for write_budget_data."""

    def _budget_data(self, loc_names, columns, values, titles):
        return [loc_names, columns, values, titles]

    def test_write_budget_data_creates_sheets(self):
        from iwfm.xls._openpyxl_backend import (
            create_workbook, write_budget_data,
        )
        loc_names = ['Region1', 'Region2']
        column_headers = [['Date', 'Value1', 'Value2'],
                          ['Date', 'Value1', 'Value2']]
        df1 = pd.DataFrame({
            'Date': [datetime(2020, 1, 1), datetime(2020, 2, 1)],
            'Value1': [100.0, 110.0],
            'Value2': [200.0, 210.0],
        })
        df2 = pd.DataFrame({
            'Date': [datetime(2020, 1, 1), datetime(2020, 2, 1)],
            'Value1': [300.0, 310.0],
            'Value2': [400.0, 410.0],
        })
        titles = [['Title 1', 'GROUNDWATER Budget', 'Subtitle 1'],
                  ['Title 2', 'GROUNDWATER Budget', 'Subtitle 2']]

        budget_data = [loc_names, column_headers, [df1, df2], titles]

        wb = create_workbook()
        write_budget_data(wb, budget_data)

        assert 'Region1' in wb.sheetnames
        assert 'Region2' in wb.sheetnames

    def test_write_budget_data_writes_titles(self):
        from iwfm.xls._openpyxl_backend import (
            create_workbook, write_budget_data, get_worksheet,
        )
        df = pd.DataFrame({
            'Date': [datetime(2020, 1, 1)],
            'Value': [100.0],
        })
        budget_data = [
            ['TestRegion'],
            [['Date', 'Value']],
            [df],
            [['Main Title', 'GROUNDWATER Budget', 'Subtitle']],
        ]
        wb = create_workbook()
        write_budget_data(wb, budget_data)

        ws = get_worksheet(wb, 'TestRegion')
        assert ws.cell(row=1, column=1).value == 'Main Title'
        assert ws.cell(row=2, column=1).value == 'GROUNDWATER Budget'
        assert ws.cell(row=3, column=1).value == 'Subtitle'

    def test_write_budget_data_writes_headers_row5(self):
        from iwfm.xls._openpyxl_backend import (
            create_workbook, write_budget_data, get_worksheet,
        )
        df = pd.DataFrame({
            'Date': [datetime(2020, 1, 1)],
            'Col1': [100.0],
            'Col2': [200.0],
        })
        budget_data = [
            ['TestRegion'],
            [['Date', 'Col1', 'Col2']],
            [df],
            [['Title1', 'Title2', 'Title3']],
        ]
        wb = create_workbook()
        write_budget_data(wb, budget_data)

        ws = get_worksheet(wb, 'TestRegion')
        assert ws.cell(row=5, column=1).value == 'Date'
        assert ws.cell(row=5, column=2).value == 'Col1'
        assert ws.cell(row=5, column=3).value == 'Col2'


# ---------------------------------------------------------------------------
# Task 6: TestLegacyFunctions
# ---------------------------------------------------------------------------

class TestLegacyFunctions:
    """Legacy function support in openpyxl backend."""

    def test_excel_init_returns_proxy(self):
        from iwfm.xls._openpyxl_backend import excel_init
        result = excel_init()
        assert result is not None

    def test_excel_new_workbook_returns_workbook(self):
        from iwfm.xls._openpyxl_backend import excel_init, excel_new_workbook
        excel = excel_init()
        wb = excel_new_workbook(excel)
        assert wb is not None
        assert hasattr(wb, 'active')

    def test_excel_kill_is_noop(self):
        from iwfm.xls._openpyxl_backend import excel_init, excel_kill
        excel = excel_init()
        excel_kill(excel)  # no exception

    def test_xl_quit_is_noop(self):
        from iwfm.xls._openpyxl_backend import excel_init, xl_quit
        excel = excel_init()
        xl_quit(excel)  # no exception


# ---------------------------------------------------------------------------
# Task 12: TestIntegration
# ---------------------------------------------------------------------------

class TestIntegration:
    """End-to-end workflow tests."""

    def test_create_save_open_modify_save(self, tmp_path):
        """Full workflow: create → save → open → modify → save → reopen."""
        from iwfm.xls import (
            create_workbook, save_workbook, open_workbook,
            close_workbook, add_worksheet, write_cells, get_worksheet,
        )
        filepath = tmp_path / "integration_test.xlsx"

        wb = create_workbook(str(filepath))
        ws = add_worksheet(wb, name='Data')
        write_cells(ws, [[1, 2], [3, 4]], start_row=1, start_col=1)
        save_workbook(wb)
        close_workbook(wb)

        wb = open_workbook(str(filepath))
        ws = get_worksheet(wb, 'Data')
        assert ws.cell(row=1, column=1).value == 1
        assert ws.cell(row=2, column=2).value == 4

        ws.cell(row=1, column=1, value=100)
        save_workbook(wb)
        close_workbook(wb)

        wb = open_workbook(str(filepath))
        ws = get_worksheet(wb, 'Data')
        assert ws.cell(row=1, column=1).value == 100
        close_workbook(wb)

    def test_backend_selection(self):
        """get_backend returns 'openpyxl' or 'win32com'."""
        from iwfm.xls import get_backend
        backend = get_backend()
        assert backend in ('openpyxl', 'win32com')

        # openpyxl available in test env → expect openpyxl
        try:
            import openpyxl  # noqa: F401
            assert backend == 'openpyxl'
        except ImportError:
            pass


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
