# hdf2xlsx_core.py
# Convert an IWFM Budget HDF5 file to an Excel workbook
# Copyright (C) 2026 University of California
# -----------------------------------------------------------------------------
# This information is free; you can redistribute it and/or modify it
# under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# This work is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# For a copy of the GNU General Public License, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301, USA.
# -----------------------------------------------------------------------------

'''Convert an IWFM Budget HDF5 file to an Excel workbook.'''

import os
import numpy as np
import h5py
import openpyxl
from openpyxl.styles import Font, Alignment

from iwfm.debug.logger_setup import logger, setup_debug_logger


def hdf2xlsx_core(hdf_file, output_file,
             budget_title='BUDGET', area_title='SUBREGION AREA',
             len_fact=1.0, len_units='FEET',
             area_fact=0.000022957, area_units='AC',
             vol_fact=0.000022957, vol_units='ACFT',
             verbose=False, debug=False):
    """
    Convert an IWFM Budget HDF5 file to an Excel workbook.

    Shared engine for all budget types; the per-type hdf2xlsx_*
    functions delegate here.

    Parameters
    ----------
    hdf_file : str
        Path to input HDF5 file
    output_file : str
        Path to output Excel file
    budget_title : str, default='BUDGET'
        Sheet title prefix, e.g. 'GROUNDWATER BUDGET' produces
        'GROUNDWATER BUDGET IN <units> FOR <location>'
    area_title : str, default='SUBREGION AREA'
        Area line prefix, e.g. 'SUBREGION AREA' produces
        'SUBREGION AREA: <area> <units>' (only written when area > 0)
    len_fact : float, default=1.0
        Length conversion factor (multiplier)
    len_units : str, default='FEET'
        Length units for output
    area_fact : float, default=0.000022957
        Area conversion factor (sq ft to acres: 0.0000229568411)
    area_units : str, default='AC'
        Area units for output (AC for acres)
    vol_fact : float, default=0.000022957
        Volume conversion factor (cu ft to acre-ft: 0.0000229568411)
    vol_units : str, default='ACFT'
        Volume units for output (ACFT for acre-feet)
    verbose : bool, default=False
        Print progress messages
    debug : bool, default=False
        Enable debug output

    Notes
    -----
    Creates an Excel workbook with:
    - Sheet1 (empty)
    - One sheet per location with budget data

    Default conversion factors:
    - Length: 1.0 (no conversion, stays in feet)
    - Area: 0.000022957 (square feet to acres, exact: 1/43560 = 0.0000229568411)
    - Volume: 0.000022957 (cubic feet to acre-feet, exact: 1/43560 = 0.0000229568411)
    """
    import iwfm

    # Configure loguru logger for debug mode
    if debug:
        setup_debug_logger()  # Auto-detects script name

    if not os.path.exists(hdf_file):
        raise FileNotFoundError(f"File '{hdf_file}' not found")

    if debug:
        logger.debug(f"Opening HDF5 file: {hdf_file}")
        logger.debug(f"Output file: {output_file}")
        logger.debug("Conversion factors:")
        logger.debug(f"  Length: {len_fact} ({len_units})")
        logger.debug(f"  Area: {area_fact} ({area_units})")
        logger.debug(f"  Volume: {vol_fact} ({vol_units})")

    # Conversion factors from model units to output units
    area_conversion = area_fact
    volume_conversion = vol_fact
    length_conversion = len_fact

    # Create Excel workbook
    wb = openpyxl.Workbook()

    # First sheet is empty (keep default Sheet)
    ws_empty = wb.active
    ws_empty.title = "Sheet1"

    with h5py.File(hdf_file, 'r') as f:
        # Get metadata
        attrs = f['Attributes'].attrs

        # Get basic info
        n_locations = attrs['nLocations']
        n_timesteps = attrs['NTimeSteps']
        n_areas = attrs['NAreas']

        # Get time step info
        start_date = attrs['TimeStep%BeginDateAndTime'].decode('utf-8')
        delta_t = attrs['TimeStep%DeltaT']
        time_unit = attrs['TimeStep%Unit'].decode('utf-8')

        if debug:
            logger.debug(f"Locations: {n_locations=}")
            logger.debug(f"Areas: {n_areas=}")
            logger.debug(f"Time steps: {n_timesteps=}")
            logger.debug(f"Start date: {start_date=}")
            logger.debug(f"Time Unit: {time_unit=}")
            logger.debug(f"Delta t: {delta_t=}")

        # Generate datetime objects for Excel
        datetime_objs = iwfm.generate_datetime_objects(start_date, n_timesteps, delta_t, time_unit)

        # Get location names and areas (if available)
        location_names = [name.decode('utf-8').strip() for name in f['Attributes/cLocationNames'][:]]
        # Areas are optional (not present in Stream budgets, for example)
        if 'Areas' in f['Attributes']:
            areas_raw = f['Attributes/Areas'][:]
            areas = areas_raw * area_conversion
        else:
            areas = [0.0] * n_locations  # Placeholder when areas not available

        # Get column header information
        full_headers = [h.decode('utf-8').strip() for h in attrs['LocationData1%cFullColumnHeaders']]
        col_types = attrs['LocationData1%iDataColumnTypes']

        # Get descriptor
        descriptor = attrs['Descriptor'].decode('utf-8')

        # Determine output unit labels
        if vol_units.upper() in ['ACFT', 'AC-FT', 'ACRE-FT', 'ACRE-FEET']:
            vol_label = 'ac.ft.'
        elif vol_units.upper() in ['AF']:
            vol_label = 'af'
        else:
            vol_label = vol_units.lower()

        if area_units.upper() in ['AC', 'ACRES']:
            area_label = 'acres'
        else:
            area_label = area_units.lower()

        # Replace placeholders in headers with actual units
        full_headers = [h.replace('@UNITAR@', area_label).replace('@UNITVL@', vol_label) for h in full_headers]

        # Process each location
        for loc_idx in range(n_locations):
            loc_name = location_names[loc_idx]
            area = areas[loc_idx]

            if debug:
                logger.debug(f"Processing location {loc_idx+1}/{n_locations}: {loc_name}")

            # Create new sheet for this location
            ws = wb.create_sheet(title=loc_name[:31])  # Excel sheet name limit is 31 chars

            # Get data for this location
            data_raw = f[loc_name][:]

            # Convert data based on column type
            # Column types: 1=volume/flow, 2=area, 3=volume stored (storage, discrepancy), 4=length
            # col_types array includes time column (index 0), so data columns start at index 1
            data = np.zeros_like(data_raw)
            for col_idx in range(data_raw.shape[1]):
                # col_types[0] is for time, col_types[1] is for first data column, etc.
                if col_idx + 1 < len(col_types):
                    col_type = col_types[col_idx + 1]
                    if col_type == 1 or col_type == 3:  # Volume (flow or storage)
                        data[:, col_idx] = data_raw[:, col_idx] * volume_conversion
                    elif col_type == 2:  # Area
                        data[:, col_idx] = data_raw[:, col_idx] * area_conversion
                    elif col_type == 4:  # Length
                        data[:, col_idx] = data_raw[:, col_idx] * length_conversion
                    else:
                        data[:, col_idx] = data_raw[:, col_idx]
                else:
                    # If no type specified, assume volume/flow
                    data[:, col_idx] = data_raw[:, col_idx] * volume_conversion

            # Write title rows
            ws['A1'] = descriptor
            ws['A1'].font = Font(bold=True)

            ws['A2'] = f"{budget_title} IN {vol_label} FOR {loc_name}"
            ws['A2'].font = Font(bold=True)

            if area > 0:
                ws['A3'] = f"{area_title}: {area:,.2f} {area_label}"

            # Write column headers (row 5)
            ws.cell(row=5, column=1, value='Time')
            # Skip first header if it's a time/date column (already written)
            headers_to_write = full_headers[1:] if full_headers and full_headers[0].lower() in ['time', 'date', 'datetime'] else full_headers
            for col_idx, header in enumerate(headers_to_write):
                ws.cell(row=5, column=col_idx + 2, value=header)

            # Make header row bold, centered, and wrap text
            for cell in ws[5]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Write data rows
            for time_idx in range(n_timesteps):
                row_num = time_idx + 6  # Data starts at row 6

                # Write datetime
                ws.cell(row=row_num, column=1, value=datetime_objs[time_idx])

                # Write data values
                for col_idx in range(data.shape[1]):
                    val = data[time_idx, col_idx]
                    ws.cell(row=row_num, column=col_idx + 2, value=val)

            # Format column A (dates): mm/dd/yyyy format, width 11.0
            ws.column_dimensions['A'].width = 11.0
            for row_num in range(6, 6 + n_timesteps):
                ws.cell(row=row_num, column=1).number_format = 'mm/dd/yyyy'

            # Format remaining columns: number format with comma separator and 2 decimals, width 12.0
            for col_idx in range(data.shape[1]):
                col_letter = openpyxl.utils.get_column_letter(col_idx + 2)
                ws.column_dimensions[col_letter].width = 12.0
                for row_num in range(6, 6 + n_timesteps):
                    ws.cell(row=row_num, column=col_idx + 2).number_format = '#,##0.00'

    # Save workbook
    wb.save(output_file)

    if verbose:
        print(f"Excel workbook written to: {output_file}")
