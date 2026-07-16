# hdf2zxlsx_gw.py
# Convert IWFM Groundwater Zone Budget HDF5 file to Excel workbook
# Copyright (C) 2020-2026 University of California
# -----------------------------------------------------------------------------
# This information is free; you can redistribute it and/or modify it
# under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# This work is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# For a copy of the GNU General Public License, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301, USA.
# -----------------------------------------------------------------------------

import openpyxl
from openpyxl.styles import Font, Alignment

from iwfm.debug.logger_setup import logger, setup_debug_logger
from iwfm.hdf5.zbud_gw_core import read_zone_definition, zbud_gw_aggregate  # noqa: F401  (read_zone_definition re-exported for compatibility)


def hdf2zxlsx_gw(hdf_file, zone_file, output_file,
                 area_fact=0.000022957, area_units='AC',
                 vol_fact=0.000022957, vol_units='ACFT',
                 verbose=False, debug=False):
    """
    Convert IWFM Groundwater Zone Budget HDF5 file to Excel workbook

    Reading and zone aggregation are shared with hdf2zbud_gw via
    zbud_gw_core.zbud_gw_aggregate(); this function only writes the Excel
    workbook (one sheet per zone).

    Parameters
    ----------
    hdf_file : str
        Path to input HDF5 zone budget file
    zone_file : str
        Path to zone definition file
    output_file : str
        Path to output Excel file (.xlsx)
    area_fact : float
        Area conversion factor (default: sq ft to acres)
    area_units : str
        Area units for output
    vol_fact : float
        Volume conversion factor (default: cu ft to acre-ft)
    vol_units : str
        Volume units for output
    verbose : bool
        Print progress messages
    debug : bool, default=False
        Enable debug output (more detailed than verbose)
    """
    import iwfm

    # Configure loguru logger for debug mode
    if debug:
        setup_debug_logger()  # Auto-detects script name

    z = zbud_gw_aggregate(hdf_file, zone_file,
                          area_fact=area_fact, vol_fact=vol_fact,
                          verbose=verbose, debug=debug)

    zone_info, zone_areas = z.zone_info, z.zone_areas
    full_headers, zone_data = z.full_headers, z.zone_data
    n_timesteps, descriptor = z.n_timesteps, z.descriptor

    # Generate datetime objects for Excel
    datetime_objs = iwfm.generate_datetime_objects(
        z.start_date, n_timesteps, z.delta_t, z.time_unit)

    # Create Excel workbook
    wb = openpyxl.Workbook()

    # First sheet is empty (keep default Sheet)
    ws_empty = wb.active
    ws_empty.title = "Sheet1"

    # Determine unit labels
    if vol_units.upper() in ['ACFT', 'AC-FT', 'ACRE-FT', 'ACRE-FEET']:
        vol_label = 'ac.ft.'
    elif vol_units.upper() in ['AF']:
        vol_label = 'af'
    else:
        vol_label = vol_units.lower()

    if area_units.upper() in ['AC', 'ACRES']:
        area_label = 'acres'
    else:
        area_label = area_units.lower()

    # Write Excel sheets
    if debug:
        logger.debug(f"Writing Excel workbook: {output_file}")

    for zone_id in sorted(zone_data.keys()):
        zone_name = zone_info.get(zone_id, f'Zone{zone_id}')
        zone_area = zone_areas.get(zone_id, 0.0)

        if debug:
            logger.debug(f"Zone {zone_id}: {zone_name}")

        # Create new sheet for this zone (max 31 chars for sheet name)
        sheet_name = f"Zone{zone_id}_{zone_name}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Write title rows
        ws['A1'] = descriptor
        ws['A1'].font = Font(bold=True)

        budget_title = f"GROUNDWATER ZONE BUDGET IN {vol_label} FOR ZONE {zone_id} ({zone_name})"
        ws['A2'] = budget_title
        ws['A2'].font = Font(bold=True)

        area_title = f"ZONE AREA: {zone_area:,.2f} {area_label}"
        ws['A3'] = area_title

        # Build column headers with IN/OUT for each component
        # Row 5: component names (merged cells for each component)
        # We'll write headers starting at row 5
        col = 1
        ws.cell(row=5, column=col, value='Time')
        col += 1

        for header in full_headers:
            # Write component name spanning 2 columns
            ws.cell(row=5, column=col, value=header)
            ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)

            # Write IN/OUT sub-headers
            ws.cell(row=6, column=col, value='IN (+)')
            ws.cell(row=6, column=col+1, value='OUT (-)')

            col += 2

        # Add Discrepancy and Absolute Storage columns
        ws.cell(row=5, column=col, value='Discrepancy')
        ws.cell(row=6, column=col, value='(=)')
        col += 1

        ws.cell(row=5, column=col, value='Absolute Storage')
        ws.cell(row=6, column=col, value='')

        # Make header rows bold, centered, and wrap text
        for cell in ws[5]:
            if cell.value:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for cell in ws[6]:
            if cell.value:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Write data rows starting at row 7
        for time_idx in range(n_timesteps):
            row_num = time_idx + 7
            col = 1

            # Write datetime
            ws.cell(row=row_num, column=col, value=datetime_objs[time_idx])
            col += 1

            total_in = 0.0
            total_out = 0.0

            # Write IN/OUT values for each component
            for comp_idx in range(len(full_headers)):
                in_val = zone_data[zone_id][comp_idx]['in'][time_idx]
                out_val = zone_data[zone_id][comp_idx]['out'][time_idx]

                ws.cell(row=row_num, column=col, value=in_val)
                ws.cell(row=row_num, column=col+1, value=out_val)
                col += 2

                total_in += in_val
                total_out += out_val

            # Calculate discrepancy
            discrepancy = total_in - total_out

            # Get absolute storage from GW Storage component
            abs_storage = 0.0
            for comp_idx, header in enumerate(full_headers):
                if 'Storage' in header or 'STORAGE' in header:
                    abs_storage = zone_data[zone_id][comp_idx]['in'][time_idx] - zone_data[zone_id][comp_idx]['out'][time_idx]
                    break

            ws.cell(row=row_num, column=col, value=discrepancy)
            col += 1
            ws.cell(row=row_num, column=col, value=abs_storage)

        # Format column A (dates): mm/dd/yyyy format, width 11.0
        ws.column_dimensions['A'].width = 11.0
        for row_num in range(7, 7 + n_timesteps):
            ws.cell(row=row_num, column=1).number_format = 'mm/dd/yyyy'

        # Format remaining columns: number format with comma separator and 2 decimals, width 12.0
        total_cols = 1 + (len(full_headers) * 2) + 2  # Time + (IN/OUT pairs) + Disc + Storage
        for col_idx in range(2, total_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 12.0
            for row_num in range(7, 7 + n_timesteps):
                ws.cell(row=row_num, column=col_idx).number_format = '#,##0.00'

    # Save workbook
    wb.save(output_file)

    if verbose:
        print(f"  Excel workbook written to: {output_file}")


if __name__ == '__main__':
    import argparse
    import iwfm.debug as idb

    parser = argparse.ArgumentParser(
        description='Convert IWFM Groundwater Zone Budget HDF5 to Excel workbook',
        formatter_class=argparse.RawDescriptionHelpFormatter)

    parser.add_argument('hdf_file', help='Input HDF5 zone budget file')
    parser.add_argument('zone_file', help='Zone definition file (.dat)')
    parser.add_argument('output_file', help='Output Excel file (.xlsx)')

    parser.add_argument('--area-fact', type=float, default=0.000022957,
                        help='Area conversion factor (default: 0.000022957 for sq ft to acres)')
    parser.add_argument('--area-units', type=str, default='AC',
                        help='Area units for output (default: AC)')

    parser.add_argument('--vol-fact', type=float, default=0.000022957,
                        help='Volume conversion factor (default: 0.000022957 for cu ft to acre-ft)')
    parser.add_argument('--vol-units', type=str, default='ACFT',
                        help='Volume units for output (default: ACFT)')

    parser.add_argument('--quiet', action='store_true', help='Suppress progress messages')

    parser.add_argument('--debug', action='store_true', help='Enable debug output')

    args = parser.parse_args()

    idb.exe_time()  # initialize timer

    hdf2zxlsx_gw(args.hdf_file, args.zone_file, args.output_file,
                 area_fact=args.area_fact, area_units=args.area_units,
                 vol_fact=args.vol_fact, vol_units=args.vol_units,
                 verbose=not args.quiet, debug=args.debug)

    idb.exe_time()  # print execution time
